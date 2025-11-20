import pandas
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from random import randint
from datetime import date
import os

STYLE_TEMPLATES = {
    "default": {
        "header_fill": PatternFill(start_color="1F4E78", fill_type="solid"),
        "header_font": Font(color="FFFFFF", bold=True),
        "header_alignment": Alignment(horizontal="center", vertical="center"),
        "row_fill": PatternFill(start_color="FFFFFF", fill_type="solid"),
        "row_font": Font(color="000000"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    },

    "light": {
        "header_fill": PatternFill(start_color="BDD7E7", fill_type="solid"),
        "header_font": Font(color="08519C", bold=True),
        "header_alignment": Alignment(horizontal="center", vertical="center"),
        "row_fill": PatternFill(start_color="EFF3FF", fill_type="solid"),
        "row_font": Font(color="3183BD"),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
    },

    "dark": {
        "header_fill": PatternFill(start_color="636363", fill_type="solid"),
        "header_font": Font(color="BDD7E7", bold=True),
        "header_alignment": Alignment(horizontal="center", vertical="center"),
        "row_fill": PatternFill(start_color="969696", fill_type="solid"),
        "row_font": Font(color="F7F7F7"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium")
        ),
    },

    "blue": {
        "header_fill": PatternFill(start_color="08519C", fill_type="solid"),
        "header_font": Font(color="BDD7E7", bold=True),
        "header_alignment": Alignment(horizontal="center", vertical="center"),
        "row_fill": PatternFill(start_color="6BAED6", fill_type="solid"),
        "row_font": Font(color="F7F7F7"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="medium", color="3182BD"), right=Side(style="medium", color="3182BD"),
            top=Side(style="medium", color="3182BD"), bottom=Side(style="medium", color="3182BD"),
        ),
    },

    "green": {
        "header_fill": PatternFill(start_color="31A354", fill_type="solid"),
        "header_font": Font(color="FFFFCC", bold=True),
        "header_alignment": Alignment(horizontal="center", vertical="center"),
        "row_fill": PatternFill(start_color="C2E699", fill_type="solid"),
        "row_font": Font(color="006837"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="medium", color="006873"), right=Side(style="medium", color="006873"),
            top=Side(style="medium", color="006873"), bottom=Side(style="medium", color="006873"),
        ),
    },

    "high_contrast": {
        "header_fill": PatternFill(start_color="000000", fill_type="solid"),
        "header_font": Font(color="FFFF00", bold=True),
        "header_alignment": Alignment(horizontal="center", vertical="center"),
        "row_fill": PatternFill(start_color="FFFFFF", fill_type="solid"),
        "row_font": Font(color="000000"),
        "alignment": Alignment(horizontal="left", vertical="center"),  
        "border": Border(
            left=Side(style="thick", color="000000"),
            right=Side(style="thick", color="000000"),
            top=Side(style="thick", color="000000"),
            bottom=Side(style="thick", color="000000")
        ),
    }
}

def fit_columns(ws):
    """
    Helper function for style_excel() to set column widths automatically based on cell contents.
    :return: None
    """
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)

        # Determine length of cell contents
        for cell in col:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass

        # Add some padding
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width


class ReportGenerator:
    """
    Generates Excel report by creating a DataFrame from inputted dictionary. Allows sorting by a specified column and
    application of templated styling.

    :param report_data: Dictionary containing the headers and rows to be turned into a DataFrame, along with indicators
    for what header the report should be sorted by (if any) and what style template should be applied (if any).
    """
    def __init__(self, report_data):
        self.dataframe = pandas.DataFrame(report_data['Rows'], columns=report_data['Headers'])
        self.sort_by = False if 'sort_by' not in report_data else report_data['sort_by']
        self.style = False if 'style' not in report_data else report_data['style']

    def needs_style(self):
        """
        Indicates whether styling needs to be applied to the report, based on the original input dictionary.
        :return:    True if styling is needed, False if not.
        """
        return True if self.style else False

    def needs_sort(self):
        """
        Indicates whether sorting needs to be applied to the report, based on original input dictionary.
        :return:    True if sorting is needed, False if not.
        """
        return True if self.sort_by else False

    def sort_report(self):
        """
        Sorts the DataFrame by the column indicated by the sort_by value of original input dictionary. The indicated
        column will be sorted in ascending order.
        :return:    None.
        """
        self.dataframe = self.dataframe.sort_values(by=self.sort_by)

    def style_report(self, file_path):
        """
        Applies predefined style template to the generated report.  Loads excel file from argument file path, applies 
        template styling to header and body rows, and saves Excel file with updates.
        :returns: original Excel file path
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Sheet1']

        template = STYLE_TEMPLATES[self.style]

        # Header row
        for cell in ws[1]:
            cell.fill = template["header_fill"]
            cell.font = template["header_font"]
            cell.alignment = template["header_alignment"]
            cell.border = template["border"]
        
        # Body rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = template["row_fill"]
                cell.font = template["row_font"]
                cell.alignment = template["header_alignment"]
                cell.border = template["border"]
        
        fit_columns(ws)

        wb.save(file_path)
        return file_path

    def generate_file_name(self):
        """
        Helper function than creates semi-random file name in [today's date]-report-randint format. Also generates
        directory 'reports' if it does not already exist.
        :return:    String containing the generated file name.
        """
        if not os.path.exists("reports"):
            os.mkdir("reports")

        file_name = f"reports/{date.today()}-report-{randint(300, 9000)}.xlsx"
        return file_name

    def generate_excel(self):
            """
            Generates an Excel file from the DataFrame, stores it in the reports directory, and returns the created files
            absolute filepath.
            :return:    String containing the file path of the created Excel file.
            """
            file_name = self.generate_file_name()

            if self.needs_sort():
                self.sort_report()

            self.dataframe.to_excel(file_name, index=False)

            if self.needs_style():
                file_name = self.style_report(file_name)

            return os.path.abspath(file_name)
